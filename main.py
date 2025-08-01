import pandas as pd
import streamlit as st
import io
import zipfile
from openpyxl import load_workbook
import openpyxl
from datetime import datetime



template_path = 'Template/Report Template.xlsx'
zip_buffer = io.BytesIO()

unified_data_model_beneficiary_object_dtypes = {
    "Sr.No": pd.StringDtype(),
    "Client Name": pd.StringDtype(),
    "TPA": pd.StringDtype(),
    "Plan": pd.StringDtype(),
    "Conversion Status": pd.StringDtype(),
    "Broker Company Name": pd.StringDtype(),
    "Broker Name": pd.StringDtype(),
    "Agent Email ID": pd.StringDtype(),
    "BRM Name ": pd.StringDtype(),

}

unified_data_model_beneficiary_dates_dtypes = [
    "policy_start_date",
    "Quote Creation Date"
]

unified_data_model_beneficiary_float_dtypes = [
    "Quoted Premium",
    "No of members Covered",
]

def types_validation(dataframe):
    """
    Validates and corrects data types in a DataFrame according to predefined rules.

    Parameters:
        dataframe (pd.DataFrame): The DataFrame to validate and correct.

    Returns:
        pd.DataFrame: The DataFrame with corrected data types.
    """
    # Define valid ranges for dates
    valid_date_range = (datetime(1900, 1, 1), datetime(2100, 12, 31))

    def correct_year_offset(date):
        """
        Corrects dates with incorrect year offsets (e.g., 2123 instead of 2023).

        Parameters:
            date (datetime): The date to correct.

        Returns:
            datetime: The corrected date.
        """
        if date is not pd.NaT:
            # If the year is far into the future, assume it is an offset error and subtract 100 years
            if date.year > datetime.now().year + 10:
                return date.replace(year=date.year - 100)
        return date

    # Step 1: Convert object columns to the specified data types
    # Iterate through object data type mappings and convert columns
    for column, dtype in unified_data_model_beneficiary_object_dtypes.items():
        if column in dataframe.columns:
            # Convert the column to the specified data type, ignoring errors
            dataframe[column] = dataframe[column].astype(dtype, errors='ignore')

    # Step 2: Convert float columns
    # Handle commas in numerical data and convert columns to numeric type
    for column in unified_data_model_beneficiary_float_dtypes:
        if column in dataframe.columns:
            # Remove commas to avoid conversion errors
            dataframe[column] = dataframe[column].astype(str).str.replace(",", "")
            # Convert to numeric, coercing invalid entries to NaN
            dataframe[column] = pd.to_numeric(dataframe[column], errors='coerce')

    # Step 3: Convert date columns
    # Validate and correct date ranges
    for column in unified_data_model_beneficiary_dates_dtypes:
        if column in dataframe.columns:
            # Convert the column to datetime, coercing invalid entries to NaT
            dataframe[column] = pd.to_datetime(dataframe[column], errors='coerce')
            # Apply range validation and correct year offsets if necessary
            dataframe[column] = dataframe[column].apply(
                lambda date: correct_year_offset(date) if valid_date_range[0] <= date <= valid_date_range[1] else pd.NaT
            )

    return dataframe



# --- File Upload ---
report = st.file_uploader("Upload Report File", type=["xlsx", "csv"])

if st.button("Generate Reports"):

    if report is not None:
        # --- Read File ---
        if report.name.endswith(".xlsx"):
            df_report = pd.read_excel(report)
            df_report = types_validation(df_report)  # Validate data types
        elif report.name.endswith(".csv"):
            df_report = pd.read_csv(report)
        else:
            st.error("Unsupported file format. Please upload an Excel or CSV file.")
            df_report = None

        if df_report is not None:
            if 'BRM' not in df_report.columns:
                st.error("Missing 'BRM' column in the report. Please check your file.")
            else:
                # --- Start Zipping Process ---
                with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED) as zip_file:

                    # --- Full Management Report ---
                    full_io = io.BytesIO()
                    wb_full = load_workbook(template_path, data_only=False)
                    ws_full = wb_full["Raw Data"]

                    # Clear old data (optional)
                    for row in ws_full.iter_rows(min_row=2, max_row=ws_full.max_row):
                        for cell in row:
                            cell.value = None

                    for r_idx, row in enumerate(df_report.values, start=2):
                        for c_idx, value in enumerate(row, start=1):
                            ws_full.cell(row=r_idx, column=c_idx, value=value)

                    wb_full.save(full_io)
                    zip_file.writestr("Management Report.xlsx", full_io.getvalue())

                    # --- Loop through each BRM ---
                    for brm in df_report['BRM'].dropna().unique():
                        st.write(f"Processing BRM: {brm}")
                        brm_data = df_report[df_report['BRM'] == brm]

                        brm_io = io.BytesIO()
                        wb_brm = load_workbook(template_path)
                        ws_brm = wb_brm["Raw Data"]

                        # Clear old data
                        for row in ws_brm.iter_rows(min_row=2, max_row=ws_brm.max_row):
                            for cell in row:
                                cell.value = None

                        for r_idx, row in enumerate(brm_data.values, start=2):
                            for c_idx, value in enumerate(row, start=1):
                                ws_brm.cell(row=r_idx, column=c_idx, value=value)
                        
                        wb_full.properties.calcPr = openpyxl.workbook.properties.CalcProperties(fullCalcOnLoad=True)

                        wb_brm.save(brm_io)
                        safe_name = f"{brm}.xlsx".replace("/", "_").replace("\\", "_")
                        zip_file.writestr(safe_name, brm_io.getvalue())

                # --- Download Button ---
                zip_buffer.seek(0)
                st.download_button(
                    label="📥 Download All Reports (ZIP)",
                    data=zip_buffer,
                    file_name="BRM_Reports.zip",
                    mime="application/zip"
                )
